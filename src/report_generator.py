from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_report(kpis, region_summary, product_summary, df, config):
    project_root = Path(__file__).resolve().parent.parent
    output_file = project_root / config["output_file"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dashboard"

    # -------------------------------------------------
    # Stiller
    # -------------------------------------------------
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78",
    )

    kpi_fill = PatternFill(
        fill_type="solid",
        start_color="D9EAF7",
        end_color="D9EAF7",
    )

    # -------------------------------------------------
    # Logo
    # -------------------------------------------------
    logo_file = project_root / config["logo_path"]

    if logo_file.exists():
        logo = Image(logo_file)
        logo.width = 120
        logo.height = 60
        sheet.add_image(logo, "G1")

    # -------------------------------------------------
    # Başlık
    # -------------------------------------------------
    sheet.merge_cells("A1:B1")

    sheet["A1"] = config["dashboard_title"]
    sheet["A1"].font = Font(
        size=16,
        bold=True,
        color="FFFFFF",
    )
    sheet["A1"].fill = header_fill
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    sheet["A2"] = config["company_name"]
    sheet["A2"].font = Font(
        size=12,
        italic=True,
    )

    # -------------------------------------------------
    # Analiz Dönemi
    # -------------------------------------------------
    if not df.empty and "Date" in df.columns:
        min_date = df["Date"].min()
        max_date = df["Date"].max()

        period_text = (
            f"Analiz Dönemi: "
            f"{min_date.strftime('%d.%m.%Y')} - "
            f"{max_date.strftime('%d.%m.%Y')}"
        )

        sheet["A3"] = period_text
        sheet["A3"].font = Font(
            size=10,
            italic=True,
        )

    # -------------------------------------------------
    # KPI
    # -------------------------------------------------
    row = 5

    for key, value in kpis.items():
        label_cell = sheet[f"A{row}"]
        value_cell = sheet[f"B{row}"]

        label_cell.value = key
        label_cell.font = Font(bold=True)
        label_cell.fill = kpi_fill
        label_cell.border = thin_border
        label_cell.alignment = Alignment(
            vertical="center"
        )

        value_cell.fill = kpi_fill
        value_cell.border = thin_border
        value_cell.alignment = Alignment(
            horizontal="right"
        )

        if isinstance(value, float):

            if "Margin" in key:
                value_cell.value = value / 100
                value_cell.number_format = "0.00%"

            else:
                value_cell.value = value
                value_cell.number_format = (
                    f'{config["currency"]}#,##0.00'
                )

        else:
            value_cell.value = value

            if "Quantity" not in key:
                value_cell.number_format = (
                    f'{config["currency"]}#,##0'
                )

        row += 1

    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 18

    # -------------------------------------------------
    # Region Summary
    # -------------------------------------------------
    region_summary = region_summary.sort_values(
        "Revenue",
        ascending=False,
    ).reset_index(drop=True)

    sheet["D2"] = "Revenue by Region"
    sheet["D2"].font = Font(
        size=14,
        bold=True,
    )

    sheet["D3"] = "Region"
    sheet["E3"] = "Revenue"

    for cell in sheet["D3:E3"][0]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )
        cell.border = thin_border

    start_row = 4

    for _, row_data in region_summary.iterrows():
        sheet[f"D{start_row}"] = row_data["Region"]
        sheet[f"E{start_row}"] = row_data["Revenue"]

        sheet[f"D{start_row}"].border = thin_border
        sheet[f"E{start_row}"].border = thin_border

        sheet[f"E{start_row}"].number_format = (
            f'{config["currency"]}#,##0'
        )

        start_row += 1

    sheet.column_dimensions["D"].width = 18
    sheet.column_dimensions["E"].width = 18

    # -------------------------------------------------
    # Region Grafiği
    # -------------------------------------------------
    chart = BarChart()

    data = Reference(
        sheet,
        min_col=5,
        min_row=3,
        max_row=start_row - 1,
    )

    categories = Reference(
        sheet,
        min_col=4,
        min_row=4,
        max_row=start_row - 1,
    )

    chart.add_data(
        data,
        titles_from_data=True,
    )

    chart.set_categories(categories)

    chart.title = "Revenue by Region"
    chart.y_axis.title = "Revenue"
    chart.x_axis.title = "Region"

    chart.style = 10
    chart.width = 12
    chart.height = 7

    sheet.add_chart(chart, "G3")

    # -------------------------------------------------
    # Top 5 Products
    # -------------------------------------------------
    product_summary = (
        product_summary
        .sort_values("Revenue", ascending=False)
        .head(5)
        .reset_index(drop=True)
    )

    sheet["D11"] = "Top 5 Products"
    sheet["D11"].font = Font(
        size=14,
        bold=True,
    )

    sheet["D12"] = "Product"
    sheet["E12"] = "Revenue"

    for cell in sheet["D12:E12"][0]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )
        cell.border = thin_border

    product_row = 13

    for _, row_data in product_summary.iterrows():
        sheet[f"D{product_row}"] = row_data["Product"]
        sheet[f"E{product_row}"] = row_data["Revenue"]

        sheet[f"D{product_row}"].border = thin_border
        sheet[f"E{product_row}"].border = thin_border

        sheet[f"E{product_row}"].number_format = (
            f'{config["currency"]}#,##0'
        )

        product_row += 1

    # -------------------------------------------------
    # Top 5 Products Grafiği
    # -------------------------------------------------
    product_chart = BarChart()

    product_data = Reference(
        sheet,
        min_col=5,
        min_row=12,
        max_row=product_row - 1,
    )

    product_categories = Reference(
        sheet,
        min_col=4,
        min_row=13,
        max_row=product_row - 1,
    )

    product_chart.add_data(
        product_data,
        titles_from_data=True,
    )

    product_chart.set_categories(
        product_categories
    )

    product_chart.title = "Top 5 Products"
    product_chart.y_axis.title = "Revenue"
    product_chart.x_axis.title = "Product"

    product_chart.style = 10
    product_chart.width = 12
    product_chart.height = 7

    sheet.add_chart(
        product_chart,
        "G18",
    )

    # -------------------------------------------------
    # Detail Data
    # -------------------------------------------------
    detail_sheet = workbook.create_sheet(
        "Detail Data"
    )

    for row_data in dataframe_to_rows(
        df,
        index=False,
        header=True,
    ):
        detail_sheet.append(row_data)

    detail_sheet.freeze_panes = "A2"
    detail_sheet.auto_filter.ref = (
        detail_sheet.dimensions
    )

    # -------------------------------------------------
    # Detail Data Başlık
    # -------------------------------------------------
    for cell in detail_sheet[1]:
        cell.font = Font(
            bold=True,
            color="FFFFFF",
        )
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )
        cell.border = thin_border

    # -------------------------------------------------
    # Detail Data Sütun Genişlikleri
    # -------------------------------------------------
    for column_cells in detail_sheet.columns:
        length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        detail_sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 3

    # -------------------------------------------------
    # Kaydet
    # -------------------------------------------------
    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook.save(output_file)